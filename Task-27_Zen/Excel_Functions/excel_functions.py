from openpyxl import load_workbook

class Excel_Functions:
    def __init__(self, excel_file_name, excel_sheet_name):
       self.file = excel_file_name
       self.sheet = excel_sheet_name


    # fetch the row count
    def row_count(self):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.max_row


    # fetch the column count
    def column_count(self):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.max_column


    # read data from excel file
    def read_data(self, row_number, column_number):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.cell(row=row_number, column=column_number).value


    # write data into excel file
    def write_data(self, row_number, column_number, data):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       sheet.cell(row=row_number, column=column_number).value = data
       workbook.save(self.file)

    # Gives details to the test_login method in Test_login page for the purpose of running pytest
    def read_excel_test_detail_util(self):
        test_details = []
        
        for row in range(2,self.row_count()+1):
            test_detail_tuple = ()
            username = self.read_data(row, 7)
            password = self.read_data(row, 8)
            exp_result = self.read_data(row, 10)
            test_detail_tuple = (username,password,row,exp_result)
            test_details.append(test_detail_tuple)

        return test_details

